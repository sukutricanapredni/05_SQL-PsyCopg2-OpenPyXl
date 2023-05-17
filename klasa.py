import psycopg2 as psycopg2
import openpyxl as openpyxl

class Ljudi:
    def __init__(self):
        self.upit=""
        self.sql_result=None
    
    def kreiraj_upit(self,upit):
        self.upit=upit

    def get_sql(self):
        try:
            con=psycopg2.connect(database='ljudi',
                                user='postgres',
                                password='itoip',
                                host='localhost',
                                port='5432')
            
            cursor=con.cursor()
            cursor.execute(self.upit)
            self.sql_result=cursor.fetchall()

        except (Exception, psycopg2.Error) as e:
           return "Error: {}".format(e)
        
        finally:
            if con:
                cursor.close()
                con.close()
    
    def export_excel(self,naziv):
        if self.sql_result!=None:
                
            wb=openpyxl.Workbook()

            ws=wb.active

            ws.title=naziv

            ws["A1"].value="JMBG"
            ws["B1"].value="Ime"
            ws["C1"].value="Prezime"
            ws["D1"].value="Godine"
            ws["E1"].value="Pol"

            for i in range(2,len(self.sql_result)+2):
                ws.cell(row=i,column=1).value=self.sql_result[i-2][0]
                ws.cell(row=i,column=2).value=self.sql_result[i-2][1]
                ws.cell(row=i,column=3).value=self.sql_result[i-2][2]
                ws.cell(row=i,column=4).value=self.sql_result[i-2][3]
                ws.cell(row=i,column=5).value=self.sql_result[i-2][4]
            

            wb.save(filename="{}.xlsx".format(naziv))
            wb.close()
            return "Excel file kreiran"
    
    def dodaj_coveka(self,jmbg,ime,prezime,godine,pol):
        try:
            con=psycopg2.connect(database='ljudi',
                                user='postgres',
                                password='itoip',
                                host='localhost',
                                port='5432')

            cursor=con.cursor()

            l='''INSERT INTO COVEK VALUES ('{}','{}','{}',{},'{}');'''.format(jmbg,ime,prezime,int(godine),pol)

            cursor.execute(l)
            con.commit()

        except (Exception,psycopg2.Error) as e:
            print("Error:",e)

        finally:
            cursor.close()
            con.close()
        
L=Ljudi()